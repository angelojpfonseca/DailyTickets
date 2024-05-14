import pandas as pd
import tkinter as tk
from tkinter import filedialog, simpledialog
from openpyxl import load_workbook

def format_excel():
    # Setup GUI for interaction
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    # Step 1: File selection via GUI
    file_path = filedialog.askopenfilename(title="Select an Excel file", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path:
        return  # No file was selected

    # Step 2: Ask user for the date
    report_date = simpledialog.askstring("Input", "Enter the report date (YYYY-MM-DD):", parent=root)
    if not report_date:
        return  # No date was entered

    # Load workbook to get sheet name
    wb = load_workbook(filename=file_path)
    first_sheet_name = wb.sheetnames[0]  # Assume the first sheet is the one to be processed

    # Step 3 to 9: Processing the Excel data
    # Load data skipping first 10 rows and using only the first 5 columns
    df = pd.read_excel(file_path, sheet_name=first_sheet_name, skiprows=10, usecols=range(5))

    # Step 5: Define data types
    df.iloc[:, 0] = pd.to_numeric(df.iloc[:, 0], errors='coerce')
    for i in range(1, 4):
        df.iloc[:, i] = pd.to_datetime(df.iloc[:, i], errors='coerce')
    df.iloc[:, 4] = df.iloc[:, 4].astype(str)

    # Step 6: Rename columns
    df.columns = ['Ticket ID', 'Requested End', 'Closed on', 'Reported on', 'Status']

    # Step 7: Add a new column
    df['Checked on'] = pd.to_datetime(report_date)

    # Step 9: Save the result with the name of the first sheet
    formatted_file_path = f"{first_sheet_name}_formatted.xlsx"
    df.to_excel(formatted_file_path, index=False)

    # Step 10: Inform the user
    print("Formatting is done and the file has been saved as:", formatted_file_path)

# Run the function to format the Excel file
format_excel()
